
import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pytz
import unicodedata
from pytz import timezone

def remover_acentos(txt):
    return ''.join(c for c in unicodedata.normalize('NFD', txt) if unicodedata.category(c) != 'Mn')

MASSAS_REGRAS = {
    "M": {
        "caracolino (box m)": "Caracolino",
        "caracolino (box kids)": "Caracolino",
        "penne (box m)": "Penne",
        "penne integral (box m)": "Penne Integral",
        "talharim (box m)": "Talharim"
    },
    "G": {
        "caracolino (box g)": "Caracolino",
        "penne (box g)": "Penne",
        "penne integral (box g)": "Penne Integral",
        "talharim (box g)": "Talharim",
        "nhoque tradicional (box g)": "Nhoque",
        "nhoque recheado de muçarela (box g)": "Nhoque Mussarela",
        "risoto de camarão": "Arroz Risoto",
        "risoto de ragu de costela": "Arroz Risoto",
        "risoto de quatro queijos": "Arroz Risoto",
        "spaguetti de abobrinha (box g)": "Espaguete Abobrinha",
        "spaguette de abobrinha (box g)": "Espaguete Abobrinha"
    }
}

MASSAS_ORDEM = {
    "G": ["Talharim", "Penne", "Penne Integral", "Caracolino", "Arroz Risoto", "Espaguete Abobrinha", "Nhoque", "Nhoque Mussarela"],
    "M": ["Talharim", "Penne", "Penne Integral", "Caracolino"]
}

SABORES_REGRAS = [
    (["quatro queijos (box g)", "nhoque quatro queijos (box g)", "risoto de quatro queijos","extra quatro queijos (box g)"], "4 Queijos", "G"),
    (["quatro queijos (box m)","extra quatro queijos (box m)"], "4 Queijos", "M"),
    (["cheddar com carne e bacon (box g)", "nhoque cheddar com carne e bacon (box g)"], ["Cheddar", "Bolonhesa"], "G"),
    (["cheddar com carne e bacon (box m)"], ["Cheddar", "Bolonhesa"], "M"),
    (["cheddar com bacon (box g)", "nhoque cheddar com bacon (box g)","extra cheddar (box g)"], "Cheddar", "G"),
    (["cheddar com bacon (box m)","cheddar com bacon (box kids)","extra cheddar (box m)"], "Cheddar", "M"),
    (["camarão rosé (box g)", "nhoque camarão rosé (box g)", "extra camarão (box g)", "risoto de camarão"], "Camarão", "G"),
    (["camarão rosé (box m)", "extra camarão (box m)"], "Camarão", "M"),
    (["extra ragu (box g)", "ragu de costela (box g)", "risoto de ragu de costela","nhoque ragu de costela (box g)"], "Ragu Costela", "G"),
    (["ragu de costela (box m)", "extra ragu (box m)"], "Ragu Costela", "M"),
    (["broccoli (box g)", "nhoque broccoli (box g)", "extra broccoli (box g)"], "Brocolis", "G"),
    (["broccoli (box m)","extra broccoli (box m)"], "Brocolis", "M"),
    (["extra presunto (box g)"],"Presunto", "G"),
    (["extra presunto (box m)"],"Presunto", "M"),
    (["parisiense (box g)", "nhoque parisiense (box g)"], ["Presunto", "Ervilha"], "G"),
    (["parisiense (box m)", "nhoque parisiense (box m)"], ["Presunto", "Ervilha"], "M"),
    (["bolonhesa (box g)", "nhoque bolonhesa (box g)", "extra carne moída (box g)"], "Bolonhesa", "G"),
    (["bolonhesa (box m)", "bolonhesa (box kids)", "nhoque bolonhesa (box m)", "extra carne moída (box m)"], "Bolonhesa", "M"),
    (["macarrão frango com requeijão cremoso (box m)"],["Frango", "Milho"],"M"),
    (["macarrão frango com requeijão cremoso (box g)"],["Frango", "Milho"],"G"),
    (["extra frango desfiado (porção 60g)"],"Frango","M"),
]

SABORES_ORDEM = {
    "G": ["Bolonhesa", "Presunto", "Ervilha", "4 Queijos", "Cheddar", "Camarão", "Ragu Costela", "Brocolis", "Frango", "Milho"],
    "M": ["Bolonhesa", "Presunto", "Ervilha", "4 Queijos", "Cheddar", "Camarão", "Ragu Costela", "Brocolis", "Frango", "Milho"]
}

def totalizar_massas(df):
    df['Itens e Opções'] = df['Itens e Opções'].fillna('').str.lower().str.strip().str.replace(r'^- ', '', regex=True)
    resultados = []
    for _, row in df.iterrows():
        item = row['Itens e Opções']
        qtd = row['Quantidade']
        for tamanho in MASSAS_REGRAS:
            if item in MASSAS_REGRAS[tamanho]:
                resultados.append({'Tipo de Massa': MASSAS_REGRAS[tamanho][item], 'Tamanho da Massa': tamanho, 'Quantidade': qtd})
                break
    zeradas = [{'Tipo de Massa': m, 'Tamanho da Massa': t, 'Quantidade': 0.0} for t, lista in MASSAS_ORDEM.items() for m in lista]
    base_df = pd.DataFrame(zeradas)
    total_df = pd.DataFrame(resultados)
    total_df['Quantidade'] = total_df['Quantidade'].astype(float)
    agregadas = total_df.groupby(['Tipo de Massa', 'Tamanho da Massa'], as_index=False)['Quantidade'].sum()
    final = base_df.merge(agregadas, how='left', on=['Tipo de Massa', 'Tamanho da Massa']).fillna(0)
    final['Quantidade'] = final['Quantidade_y']
    final = final[[c for c in final.columns if not c.endswith('_x') and not c.endswith('_y')]]
    final['ordem'] = final.apply(lambda r: MASSAS_ORDEM[r['Tamanho da Massa']].index(r['Tipo de Massa']) if r['Tipo de Massa'] in MASSAS_ORDEM[r['Tamanho da Massa']] else 999, axis=1)
    return final.sort_values(by=['Tamanho da Massa', 'ordem']).drop(columns='ordem')

def totalizar_sabores(df):
    df['Itens e Opções'] = df['Itens e Opções'].fillna('').str.lower().str.strip().str.replace(r'^- ', '', regex=True)
    resultados = []
    for item, qtd in zip(df['Itens e Opções'], df['Quantidade']):
        for entradas, sabores, tamanho in SABORES_REGRAS:
            if item in entradas:
                if isinstance(sabores, list):
                    for sabor in sabores:
                        resultados.append({'Sabor': sabor, 'Tamanho': tamanho, 'Quantidade': qtd})
                else:
                    resultados.append({'Sabor': sabores, 'Tamanho': tamanho, 'Quantidade': qtd})
                break
    zeradas = [{'Sabor': s, 'Tamanho': t, 'Quantidade': 0.0} for t, lista in SABORES_ORDEM.items() for s in lista]
    base_df = pd.DataFrame(zeradas)
    total_df = pd.DataFrame(resultados)
    total_df['Quantidade'] = total_df['Quantidade'].astype(float)
    agregadas = total_df.groupby(['Sabor', 'Tamanho'], as_index=False)['Quantidade'].sum()
    final = base_df.merge(agregadas, how='left', on=['Sabor', 'Tamanho']).fillna(0)
    final['Quantidade'] = final['Quantidade_y']
    final = final[[c for c in final.columns if not c.endswith('_x') and not c.endswith('_y')]]
    final['ordem'] = final.apply(lambda r: SABORES_ORDEM[r['Tamanho']].index(r['Sabor']) if r['Sabor'] in SABORES_ORDEM[r['Tamanho']] else 999, axis=1)
    return final.sort_values(by=['Tamanho', 'ordem']).drop(columns='ordem')

def extrair_diversos(df):
    usados = set()
    for tam in MASSAS_REGRAS.values():
        usados.update(tam.keys())
    for entradas, *_ in SABORES_REGRAS:
        usados.update(entradas)
    df['Itens e Opções'] = df['Itens e Opções'].fillna('').str.lower().str.strip().str.replace(r'^- ', '', regex=True)
    diversos = df[~df['Itens e Opções'].isin(usados)].copy()
    diversos = diversos[diversos['Itens e Opções'].str.strip() != '']
    diversos['Item'] = diversos['Itens e Opções'].str.strip()
    total = diversos.groupby('Item', as_index=False)['Quantidade'].sum()
    return total.sort_values('Item')

def main(caminho_planilha):
    df_bruto = pd.read_excel(caminho_planilha, skiprows=3)
    df_bruto = df_bruto.dropna(subset=['Itens e Opções', 'Quantidade'])

    # Lendo datas diretamente das células A2 e B2
    df_datas = pd.read_excel(caminho_planilha, nrows=2, header=None)
    data_inicial = pd.to_datetime(df_datas.iloc[1, 0], errors='coerce')
    data_final = pd.to_datetime(df_datas.iloc[1, 1], errors='coerce')

    hora_brasil = datetime.now(timezone("America/Sao_Paulo"))
    hora_str = hora_brasil.strftime("%H-%M-%S")
    nome_saida = f"italin-de-{data_inicial.strftime('%d-%m-%Y')}-a-{data_final.strftime('%d-%m-%Y')}-{hora_str}.xlsx"

    buffer = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    abas = [
        ("Massas", totalizar_massas(df_bruto).rename(columns={'Tipo de Massa': 'Massa', 'Tamanho da Massa': 'Tamanho'})),
        ("Sabores", totalizar_sabores(df_bruto)),
        ("Diversos", extrair_diversos(df_bruto))
    ]

    for nome_aba, df in abas:
        ws = wb.create_sheet(title=nome_aba)
        colunas = list(df.columns)

        ws.append(colunas)
        
        linha_excel = 2
        for row in df.itertuples(index=False):
            ws.append(row)
            linha_excel += 1
            for col in range(1, len(colunas)+1):
                ws.cell(row=linha_excel, column=col).value = ""
            linha_excel += 1

        header_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
        header_font = Font(bold=True)
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{ws.max_row}"
        for col_idx in range(1, len(df.columns) + 1):
            max_len = max(len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 5

    wb.save(buffer)
    buffer.seek(0)
    return nome_saida, buffer

# === Streamlit UI ===

st.set_page_config(page_title="Totalizador de Massas e Sabores")
st.title("🍝 Totalizador de Massas e Sabores")
uploaded = st.file_uploader("Envie sua planilha de entrada (.xlsx)", type=["xlsx"])

if uploaded:
    nome_saida, resultado = main(uploaded)
    st.success("✅ Planilha gerada com sucesso!")
    st.download_button("📥 Baixar planilha de resultados", resultado, file_name=nome_saida, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
